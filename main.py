# !/usr/bin/env python
# -*- coding: utf-8 -*-
# Created by Luis Fuentes  
# To run: python main.py

from __future__ import annotations
from typing import Optional
from etl.DBConn import SQLServerConn, OracleConnection
from etl.GeneralFunctions import (get_files_in_directory, check_if_list_Null, clean_string, is_int, 
                                  to_gcs_bucket, delete_processed_file, sub, download_blob)
from sqlalchemy.sql import text
from openpyxl import load_workbook #xlsx files
from openpyxl.styles import Font, Fill, PatternFill
from pathlib import Path
from os.path import basename, exists
from os import getcwd, makedirs
import sys

if __name__ == "__main__":

    try:

        files_directory_path = Path(getcwd(),'data')

        if not exists(files_directory_path):
            makedirs('data')
        
        files_to_process = sub('microstrategyit','DataEngineering-appusma206_apps_sub')

        if len(files_to_process) != 0:
            for _file in files_to_process:
                download_blob(_file, Path(files_directory_path,basename(_file)))

        list_of_files: list = get_files_in_directory(files_directory_path)
        
        if not check_if_list_Null(list_of_files):
            for file_name in list_of_files:

                # This is the full path for the file in the system
                file_final_path: str = "{files_directory_path}\\{file_name}".format(files_directory_path=files_directory_path, file_name = file_name)
                
                # Opening the file
                wb = load_workbook(filename=file_final_path, read_only=False)
                ws = wb.worksheets[0]

                # Create connections for DBS
                sql_s = SQLServerConn()
                sql_o = OracleConnection()

                # This is a hardcoded value
                maxShipId: int = 2910597

                # Creating a dictionary with column names and values starting at 1                
                switcher: dict = {}
                for column in range(1, ws.max_column+1):
                    cell_value: str = ws.cell(1, column).value
                    switcher[cell_value.strip().upper()] = column

                # Spreed sheed must contain this values
                registry = {
                    'CARRIER_CODE': switcher['CARRIER_CODE'],
                    'ACCT NBR': switcher['ACCT NBR'],
                    'CURRENCY': switcher['CURRENCY'],
                    'BOL': switcher['BOL'],
                    'PO': switcher['PO'],
                    'INVOICE NUMBER': switcher['INVOICE NUMBER'],
                    'HANDLING UNITS': switcher['HANDLING UNITS'],
                    'SHIP DATE': switcher['SHIP DATE'],
                    'ALLOCATE AMOUNT': switcher['ALLOCATE AMOUNT'],
                    'TRACKING NUMBER': switcher['TRACKING NUMBER']
                }

                # This are the values to add to the Spreedsheet
                values_to_add = {
                    80: 'Clarify Case Number',
                    81: 'Clarify Case Type',
                    82: 'GEO Code',
                    83: 'Amount',
                    84: 'Allocation Method'
                }

                del switcher

                if not all(registry.values()):
                    # One or many of the columns are missing
                    sys.exit()
                else:
                    print("Processing Begin")
                    # Set up Headers
                    for column in range(80, 85):
                        xcel = ws.cell(row = 1, column = column)
                        xcel.value = values_to_add[column]
                        xcel.font = Font(bold=True)
                        xcel.fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
                    
                    # Loop through rows
                    for row in range(2, ws.max_row+1):

                        print("Row Number:"+str(row))
                        
                        # Get attributes from worksheet and clean them
                        bol: str = clean_string(str(ws.cell(row, registry['BOL']).value))
                        po: str = clean_string(str(ws.cell(row, registry['PO']).value))
                        tracking_num: str = clean_string(str(ws.cell(row, registry['TRACKING NUMBER']).value))
                        carrier_code: str = clean_string(str(ws.cell(row, registry['CARRIER_CODE']).value))
                        acct_nbr: str = clean_string(str(ws.cell(row, registry['ACCT NBR']).value))
                        amount: float = float(clean_string(str(ws.cell(row, registry['ALLOCATE AMOUNT']).value)))
                        result = None

                        values = {
                            'dtlTicket': None, #80
                            'dtlTicketType':  None, #81
                            'dtlGEOCd': None, #82
                            'dtlAllocMethod': None, #84 
                        }

                        for counter in range(1, 9):
                            if counter == 1:
                                if carrier_code == "PURCO" and acct_nbr == '8557975':
                                    s = text("""
                                        SELECT xref_value 
                                        FROM xref_table 
                                        WHERE xref_type = 'PUROACCT2GEO' 
                                         AND keychar1 = :acct_nbr
                                    """)
                                    result = sql_s.sql_to_data(sqlText = s, acct_nbr = acct_nbr) # -> [{'xref_value': 'DEPOT'}] 
                                    if len(result) != 0:
                                        values['dtlTicket'] = ""
                                        values['dtlTicketType'] = ""
                                        values['dtlGEOCd'] = result[0]['xref_value'].strip()
                                        values['dtlAllocMethod'] =  "Purolator Account GEO"
                                        break
                            if counter == 2:
                                if tracking_num != "":
                                    dict_values = {'tracking_num': tracking_num}
                                    s2 = """SELECT d.objid, d.x_inbound_waybill, h.header_case_no, c.x_geo_code , c.x_service_type 
                                        FROM t_clar_demand_dtl d, t_clar_demand_hdr h, t_clar_case c 
                                        WHERE d.demand_dtl2demand_hdr = h.objid 
                                        AND h.header_case_no = c.s_id_number 
                                        AND d.x_inbound_waybill = :tracking_num"""
                                    result = sql_o.sql_to_data(s2, dict_values=dict_values)
                                    if len(result) != 0:
                                        values['dtlTicket'] = clean_string(str(result[0][2]))
                                        values['dtlTicketType'] = clean_string(str(result[0][4]))
                                        values['dtlGEOCd'] = clean_string(str(result[0][3]))
                                        values['dtlAllocMethod'] =  "Inbound Waybill " + tracking_num
                                        break
                            if counter == 3:
                                if tracking_num != "":
                                    dict_values = {'tracking_num': tracking_num}
                                    s2 = """SELECT d.objid, d.x_inbound_waybill, h.header_case_no, c.x_geo_code , c.x_service_type
                                            FROM t_clar_demand_dtl d, t_clar_demand_hdr h, t_clar_case c
                                            WHERE d.demand_dtl2demand_hdr = h.objid
                                            AND h.header_case_no = c.s_id_number
                                            AND d.waybill = :tracking_num"""
                                    result = sql_o.sql_to_data(s2, dict_values=dict_values)
                                    if len(result) != 0:
                                        values['dtlTicket'] = clean_string(str(result[0][2]))
                                        values['dtlTicketType'] = clean_string(str(result[0][4]))
                                        values['dtlGEOCd'] = clean_string(str(result[0][3]))
                                        values['dtlAllocMethod'] =  "Waybill " + tracking_num
                                        break
                            if counter == 4:
                                if tracking_num != "":
                                    dict_values = {'tracking_num': tracking_num}
                                    s2 = """SELECT d.objid, d.x_inbound_waybill, h.header_case_no, c.x_geo_code , c.x_service_type 
                                            FROM t_clar_part_used u, t_clar_demand_dtl d, t_clar_demand_hdr h, t_clar_case c
                                            WHERE u.x_waybill_no = :tracking_num
                                            AND u.part_used2demand_dtl = d.objid
                                            AND d.demand_dtl2demand_hdr = h.objid
                                            AND h.header_case_no = c.s_id_number"""
                                    result = sql_o.sql_to_data(s2, dict_values=dict_values)
                                    if len(result) != 0:
                                        values['dtlTicket'] = clean_string(str(result[0][2]))
                                        values['dtlTicketType'] = clean_string(str(result[0][4]))
                                        values['dtlGEOCd'] = clean_string(str(result[0][3]))
                                        values['dtlAllocMethod'] =  "Usage Waybill " + tracking_num
                                        break
                            if counter == 5:
                                if bol != '' and bol[0] == 'P':
                                    dict_values = {'bol': bol}
                                    s2 = """SELECT d.objid, d.x_inbound_waybill, h.header_case_no, c.x_geo_code, c.x_service_type
                                            FROM t_clar_demand_dtl d, t_clar_demand_hdr h, t_clar_case c
                                            WHERE d.demand_dtl2demand_hdr = h.objid
                                            AND h.header_case_no = c.s_id_number
                                            AND d.detail_number = :bol"""
                                    result = sql_o.sql_to_data(s2, dict_values=dict_values)
                                    if len(result) != 0:
                                        values['dtlTicket'] = clean_string(str(result[0][2]))
                                        values['dtlTicketType'] = clean_string(str(result[0][4]))
                                        values['dtlGEOCd'] = clean_string(str(result[0][3]))
                                        values['dtlAllocMethod'] =  "Amdocs Part Demand " + bol
                                        break                            
                            if counter == 6:
                                if bol != "" and bol[0] == 'C' and bol[7:8] == '-':
                                    dict_values = {'bol': bol}
                                    s2 = """SELECT s_id_number, x_geo_code, x_service_type
                                            FROM t_clar_case c
                                            WHERE s_id_number = :bol"""
                                    result = sql_o.sql_to_data(s2, dict_values=dict_values)
                                    if len(result) != 0:
                                        values['dtlTicket'] = clean_string(str(result[0][0]))
                                        values['dtlTicketType'] = clean_string(str(result[0][2]))
                                        values['dtlGEOCd'] = clean_string(str(result[0][1]))
                                        values['dtlAllocMethod'] =  "DW Case " + bol
                                        break    
                            if counter == 7:
                                if po != '' and len(po) == 3 and is_int(po):
                                    dict_values = {'po': po}
                                    s2 = """SELECT l.s_title as cdtype, e.s_title as cdval, e.rank, e.state, e.Description
                                            FROM dw.t_clar_gbst_lst l, dw.t_clar_gbst_elm e
                                            WHERE e.GBST_ELM2GBST_LST = l.objid
                                            AND l.S_TITLE like 'GEOGRAPHY CODE'
                                            AND e.s_title = :po
                                            AND e.state = '0'"""
                                    result = sql_o.sql_to_data(s2, dict_values=dict_values)
                                    if len(result) != 0:
                                        values['dtlTicket'] = ""
                                        values['dtlTicketType'] = ""
                                        values['dtlGEOCd'] = po
                                        values['dtlAllocMethod'] =  "Cust Ref 3 GEO"
                                        break
                            if counter == 8:
                                if carrier_code == "PURCO":
                                    s = text("""
                                        SELECT xref_value
                                        FROM xref_table
                                        WHERE xref_type = 'PUROACCT2GEO'
                                        AND keychar1 = :acct_nbr
                                    """)
                                    result = sql_s.sql_to_data(sqlText = s, acct_nbr = acct_nbr) # -> [{'xref_value': 'DEPOT'}] 
                                    if len(result) != 0:
                                        values['dtlTicket'] = ""
                                        values['dtlTicketType'] = ""
                                        values['dtlGEOCd'] = result[0]['xref_value'].strip()
                                        values['dtlAllocMethod'] =  "Purolator Account GEO"
                                        break                                    

                        ws.cell(row = row, column = 80).value = values['dtlTicket']
                        ws.cell(row = row, column = 81).value = values['dtlTicketType']
                        ws.cell(row = row, column = 82).value = values['dtlGEOCd']
                        ws.cell(row = row, column = 83).value = amount
                        ws.cell(row = row, column = 84).value = values['dtlAllocMethod']

                        counter += 1

                wb.save(file_final_path)

                # Upload the file to bucket
                to_gcs_bucket(file_name = file_name, file_final_path = file_final_path)

                # Delete file from server
                delete_processed_file(file_final_path)
   
    except Exception as e:
        print(e)
        print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
        sys.exit()